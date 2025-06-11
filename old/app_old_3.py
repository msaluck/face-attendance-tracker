import os
import pickle
import datetime
import csv
import tkinter as tk
import shutil
from tkinter import (
    messagebox,
    filedialog,
    simpledialog,
    Toplevel,
    Text,
    Scrollbar,
    Label,
    Entry,
    Button,
)
from openpyxl import Workbook, load_workbook
import cv2
import face_recognition
import numpy as np


# --- SETUP ---
data_dir = "data"
encoding_file = os.path.join(data_dir, "encodings.pkl")
attendance_csv = os.path.join(data_dir, "attendance.csv")
attendance_xlsx = os.path.join(data_dir, "attendance.xlsx")
os.makedirs(data_dir, exist_ok=True)

# Global variables
known_encodings = []
# --- CHANGE: Store a list of dictionaries instead of just names ---
known_person_data = []


# --- CORE FUNCTIONS ---
def save_encodings():
    """Saves the known face encodings and person data to a pickle file."""
    with open(encoding_file, "wb") as f:
        # --- CHANGE: Save person data dictionary list ---
        pickle.dump((known_encodings, known_person_data), f)


def load_encodings():
    """Loads face encodings and person data from a pickle file."""
    global known_encodings, known_person_data
    if os.path.exists(encoding_file):
        with open(encoding_file, "rb") as f:
            # --- CHANGE: Load person data dictionary list ---
            known_encodings, known_person_data = pickle.load(f)
    else:
        known_encodings = []
        known_person_data = []


def mark_attendance(person_data):
    """Marks attendance with NIS, Name, and Class."""
    now = datetime.datetime.now()
    date_str = now.strftime("%Y-%m-%d")
    time_str = now.strftime("%H:%M:%S")
    file_format = selected_format.get()

    # --- CHANGE: Prepare row with new data ---
    row_data = [
        person_data["nis"],
        person_data["name"],
        person_data["class"],
        date_str,
        time_str,
    ]
    header = ["NIS", "Name", "Class", "Date", "Time"]

    if file_format == "xlsx":
        if os.path.exists(attendance_xlsx):
            wb = load_workbook(attendance_xlsx)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(header)

        ws.append(row_data)
        wb.save(attendance_xlsx)

    elif file_format == "csv":
        with open(attendance_csv, "a", newline="") as f:
            writer = csv.writer(f)
            if os.stat(attendance_csv).st_size == 0:
                writer.writerow(header)
            writer.writerow(row_data)


def register_face():
    """--- CHANGE: Opens a form to get student details before capturing face. ---"""

    # Create a new window for the registration form
    form_window = Toplevel(root)
    form_window.title("Register New Student")
    form_window.geometry("300x200")
    form_window.grab_set()  # Modal window

    Label(form_window, text="Name:").pack(pady=(10, 0))
    name_entry = Entry(form_window, width=30)
    name_entry.pack()

    Label(form_window, text="NIS (Student ID):").pack(pady=(10, 0))
    nis_entry = Entry(form_window, width=30)
    nis_entry.pack()

    Label(form_window, text="Class:").pack(pady=(10, 0))
    class_entry = Entry(form_window, width=30)
    class_entry.pack()

    def submit_and_capture():
        name = name_entry.get().strip()
        nis = nis_entry.get().strip()
        class_ = class_entry.get().strip()  # Use class_ to avoid keyword conflict

        if not name or not nis or not class_:
            messagebox.showerror(
                "Error", "All fields are required.", parent=form_window
            )
            return

        form_window.destroy()  # Close form before opening camera

        # Proceed with face capture
        capture_face_for_registration({"name": name, "nis": nis, "class": class_})

    Button(
        form_window, text="Submit and Capture Face", command=submit_and_capture
    ).pack(pady=20)


def capture_face_for_registration(person_data):
    """Handles the face capture process after getting user details."""
    cap = cv2.VideoCapture(0)
    if not cap.isOpened():
        messagebox.showerror("Error", "Could not open webcam.")
        return

    messagebox.showinfo(
        "Capture",
        f"Capturing for {person_data['name']}.\nLook at the camera and press 's' to capture, or 'q' to cancel.",
    )

    while True:
        ret, frame = cap.read()
        if not ret:
            break
        cv2.imshow("Register Face - Press 's' to save, 'q' to quit", frame)
        key = cv2.waitKey(1) & 0xFF

        if key == ord("s"):
            rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            boxes = face_recognition.face_locations(rgb_frame)
            encodings = face_recognition.face_encodings(rgb_frame, boxes)

            if encodings:
                known_encodings.append(encodings[0])
                # --- CHANGE: Append the whole dictionary ---
                known_person_data.append(person_data)
                save_encodings()
                messagebox.showinfo(
                    "Success", f"Face registered for {person_data['name']}!"
                )
            else:
                messagebox.showerror("Error", "No face detected. Please try again.")
            break
        elif key == ord("q"):
            break

    cap.release()
    cv2.destroyAllWindows()


def recognize_face():
    """Recognizes faces using the webcam and marks attendance."""
    load_encodings()
    cap = cv2.VideoCapture(0)
    if not cap.isOpened():
        messagebox.showerror("Error", "Could not open webcam.")
        return

    # --- CHANGE: Use NIS for duplicate checking ---
    todays_attendance_nis = set()
    file_format = selected_format.get()
    date_str = datetime.datetime.now().strftime("%Y-%m-%d")

    try:
        if file_format == "xlsx" and os.path.exists(attendance_xlsx):
            wb = load_workbook(attendance_xlsx)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                # Check if row has enough columns before accessing indices
                if len(row) > 3 and row[3] == date_str:
                    todays_attendance_nis.add(row[0])  # Column 0 is NIS
        elif file_format == "csv" and os.path.exists(attendance_csv):
            with open(attendance_csv, "r") as f:
                reader = csv.reader(f)
                next(reader, None)  # Skip header
                for row in reader:
                    if len(row) > 3 and row[3] == date_str:
                        todays_attendance_nis.add(row[0])  # Column 0 is NIS
    except Exception as e:
        messagebox.showerror("Error", f"Could not read attendance file: {e}")

    while True:
        ret, frame = cap.read()
        if not ret:
            break

        small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
        rgb_small_frame = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB)
        face_locations = face_recognition.face_locations(rgb_small_frame)
        face_encodings = face_recognition.face_encodings(
            rgb_small_frame, face_locations
        )
        tolerance = tolerance_var.get()

        for (top, right, bottom, left), face_encoding in zip(
            face_locations, face_encodings
        ):
            matches = face_recognition.compare_faces(
                known_encodings, face_encoding, tolerance=tolerance
            )
            display_name = "Unknown"
            confidence_str = ""

            if True in matches:
                face_distances = face_recognition.face_distance(
                    known_encodings, face_encoding
                )
                best_match_index = np.argmin(face_distances)
                if matches[best_match_index]:
                    # --- CHANGE: Get full person data ---
                    person_data = known_person_data[best_match_index]
                    display_name = person_data["name"]

                    confidence = 1 - face_distances[best_match_index]
                    confidence_str = f"{confidence:.0%}"

                    # --- CHANGE: Check against NIS for duplicates ---
                    if person_data["nis"] not in todays_attendance_nis:
                        mark_attendance(person_data)
                        todays_attendance_nis.add(person_data["nis"])
                        print(
                            f"Attendance marked for {display_name} ({person_data['nis']})"
                        )

            top, right, bottom, left = top * 4, right * 4, bottom * 4, left * 4
            cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 0), 2)
            label = f"{display_name} {confidence_str}"
            cv2.putText(
                frame,
                label,
                (left, top - 10),
                cv2.FONT_HERSHEY_SIMPLEX,
                0.7,
                (0, 255, 0),
                2,
            )

        cv2.imshow("Face Recognition - Press 'q' to Exit", frame)
        if cv2.waitKey(1) & 0xFF == ord("q"):
            break

    cap.release()
    cv2.destroyAllWindows()


def show_logs():
    """Shows logs in a new, scrollable window from the correct file."""
    log_window = Toplevel(root)
    log_window.title("Attendance Logs")
    log_window.geometry("600x400")

    text_widget = Text(log_window, wrap="word", font=("Courier New", 10))
    scrollbar = Scrollbar(log_window, command=text_widget.yview)
    text_widget.config(yscrollcommand=scrollbar.set)
    scrollbar.pack(side="right", fill="y")
    text_widget.pack(side="left", fill="both", expand=True)

    file_format = selected_format.get()
    log_file_path = attendance_xlsx if file_format == "xlsx" else attendance_csv

    if not os.path.exists(log_file_path):
        text_widget.insert("end", "No attendance records found.")
        text_widget.config(state="disabled")
        return

    try:
        content = []
        if file_format == "xlsx":
            wb = load_workbook(log_file_path)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                content.append([str(cell) for cell in row])
        else:
            with open(log_file_path, "r") as f:
                reader = csv.reader(f)
                for row in reader:
                    content.append(row)

        # Format content for display
        if content:
            # Simple column formatting
            formatted_content = ""
            for row in content:
                # Pad each item to align columns
                formatted_row = (
                    f"{row[0]:<12}{row[1]:<25}{row[2]:<12}{row[3]:<12}{row[4]:<10}\n"
                )
                formatted_content += formatted_row
            text_widget.insert("end", formatted_content)

    except Exception as e:
        text_widget.insert("end", f"Error reading log file: {e}")

    text_widget.config(state="disabled")


def export_logs():
    """Exports logs to a user-selected location."""
    file_format = selected_format.get()
    source_file = attendance_xlsx if file_format == "xlsx" else attendance_csv

    if not os.path.exists(source_file):
        messagebox.showwarning("Warning", "No attendance file to export.")
        return

    export_path = filedialog.asksaveasfilename(
        defaultextension=f".{file_format}",
        filetypes=[(f"{file_format.upper()} Files", f"*.{file_format}")],
    )

    if export_path:
        try:
            shutil.copy(source_file, export_path)
            messagebox.showinfo(
                "Export", f"Logs successfully exported to {export_path}"
            )
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export file: {e}")


# --- GUI SETUP ---
root = tk.Tk()
root.title("Face Recognition Attendance")
root.geometry("450x450")
root.configure(bg="#f0f0f0")

main_frame = tk.Frame(root, bg="#f0f0f0", padx=20, pady=20)
main_frame.pack(expand=True, fill="both")

title_label = tk.Label(
    main_frame, text="Attendance System", font=("Helvetica", 16, "bold"), bg="#f0f0f0"
)
title_label.pack(pady=(0, 20))

btn_register = tk.Button(
    main_frame, text="Register New Student", command=register_face, width=30, height=2
)
btn_register.pack(pady=5)

btn_recognize = tk.Button(
    main_frame,
    text="Start Recognition & Attendance",
    command=recognize_face,
    width=30,
    height=2,
)
btn_recognize.pack(pady=5)

btn_logs = tk.Button(
    main_frame, text="Show Attendance Logs", command=show_logs, width=30, height=2
)
btn_logs.pack(pady=5)

btn_export = tk.Button(
    main_frame, text="Export Logs", command=export_logs, width=30, height=2
)
btn_export.pack(pady=5)

format_frame = tk.Frame(main_frame, bg="#f0f0f0")
format_frame.pack(pady=10)
format_label = tk.Label(format_frame, text="Log Format:", bg="#f0f0f0")
format_label.pack(side="left", padx=(0, 10))
selected_format = tk.StringVar(value="xlsx")
format_menu = tk.OptionMenu(format_frame, selected_format, "xlsx", "csv")
format_menu.config(width=8)
format_menu.pack(side="left")

tolerance_frame = tk.Frame(main_frame, bg="#f0f0f0")
tolerance_frame.pack(pady=10, fill="x")
tolerance_label = tk.Label(
    tolerance_frame, text="Recognition Tolerance (Stricter <-> Looser)", bg="#f0f0f0"
)
tolerance_label.pack()
tolerance_var = tk.DoubleVar(value=0.6)
tolerance_slider = tk.Scale(
    tolerance_frame,
    from_=0.4,
    to=0.7,
    resolution=0.05,
    orient="horizontal",
    variable=tolerance_var,
    bg="#f0f0f0",
)
tolerance_slider.pack(fill="x", expand=True)

load_encodings()
root.mainloop()
