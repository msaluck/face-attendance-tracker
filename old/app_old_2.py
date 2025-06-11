import os
import pickle
import datetime
import csv
import tkinter as tk
import shutil
from tkinter import messagebox, filedialog, simpledialog
from openpyxl import Workbook, load_workbook
import cv2
import face_recognition
import numpy as np


data_dir = "data"
encoding_file = os.path.join(data_dir, "encodings.pkl")
attendance_csv = os.path.join(data_dir, "attendance.csv")
attendance_xlsx = os.path.join(data_dir, "attendance.xlsx")
os.makedirs(data_dir, exist_ok=True)

known_encodings = []
known_names = []


def save_encodings():
    with open(encoding_file, "wb") as f:
        pickle.dump((known_encodings, known_names), f)


def load_encodings():
    global known_encodings, known_names
    if os.path.exists(encoding_file):
        with open(encoding_file, "rb") as f:
            known_encodings, known_names = pickle.load(f)
    else:
        known_encodings = []
        known_names = []


def mark_attendance(name):
    now = datetime.datetime.now()
    date_str = now.strftime("%Y-%m-%d")
    time_str = now.strftime("%H:%M:%S")

    file_format = selected_format.get()  # Get user choice

    if file_format == "xlsx":
        # Save to Excel file
        if os.path.exists(attendance_xlsx):
            wb = load_workbook(attendance_xlsx)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(["Name", "Date", "Time"])  # Header row

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == name and row[1] == date_str:
                return  # Prevent duplicate attendance

        ws.append([name, date_str, time_str])
        wb.save(attendance_xlsx)

    elif file_format == "csv":
        # Save to CSV file
        already_logged = set()
        if os.path.exists(attendance_csv):
            with open(attendance_csv, "r") as f:
                reader = csv.reader(f)
                next(reader, None)
                for row in reader:
                    if len(row) >= 3 and row[0] == name and row[1] == date_str:
                        already_logged.add(name)

        if name not in already_logged:
            with open(attendance_csv, "a", newline="") as f:
                writer = csv.writer(f)
                if os.stat(attendance_csv).st_size == 0:
                    writer.writerow(["Name", "Date", "Time"])
                writer.writerow([name, date_str, time_str])


def register_face():
    name = simpledialog.askstring("Register Face", "Enter Name:")
    if not name:
        return

    cap = cv2.VideoCapture(0)
    messagebox.showinfo("Capture", "Press 's' to capture the face, 'q' to cancel.")

    while True:
        ret, frame = cap.read()
        if not ret:
            break
        cv2.imshow("Register Face", frame)
        key = cv2.waitKey(1) & 0xFF

        if key == ord("s"):
            rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            boxes = face_recognition.face_locations(rgb)
            encodings = face_recognition.face_encodings(rgb, boxes)

            if encodings:
                known_encodings.append(encodings[0])
                known_names.append(name)
                save_encodings()
                messagebox.showinfo("Success", f"Face registered for {name}!")
            else:
                messagebox.showerror("Error", "No face detected. Try again.")
            break
        elif key == ord("q"):
            break

    cap.release()
    cv2.destroyAllWindows()


def recognize_face():
    load_encodings()
    cap = cv2.VideoCapture(0)
    recognized_names = set()

    while True:
        ret, frame = cap.read()
        if not ret:
            break

        small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
        # rgb_small_frame = small_frame[:, :, ::-1]
        rgb_small_frame = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB)

        face_locations = face_recognition.face_locations(rgb_small_frame)
        face_encodings = face_recognition.face_encodings(
            rgb_small_frame, face_locations
        )

        for (top, right, bottom, left), face_encoding in zip(
            face_locations, face_encodings
        ):
            matches = face_recognition.compare_faces(known_encodings, face_encoding)
            name = "Unknown"
            confidence = 0.0

            if True in matches:
                face_distances = face_recognition.face_distance(
                    known_encodings, face_encoding
                )
                best_match_index = np.argmin(face_distances)
                if matches[best_match_index]:
                    name = known_names[best_match_index]
                    confidence = (1 - face_distances[best_match_index]) * 100

                    if name not in recognized_names:
                        recognized_names.add(name)
                        mark_attendance(name)

            top *= 4
            right *= 4
            bottom *= 4
            left *= 4

            cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 0), 2)
            label = f"{name} ({confidence:.2f}%)" if name != "Unknown" else name
            cv2.putText(
                frame,
                label,
                (left, top - 10),
                cv2.FONT_HERSHEY_SIMPLEX,
                0.9,
                (0, 255, 0),
                2,
            )

        cv2.imshow("Face Recognition - Press Q to Exit", frame)
        if cv2.waitKey(1) & 0xFF == ord("q"):
            break

    cap.release()
    cv2.destroyAllWindows()


def show_logs():
    if not os.path.exists(attendance_csv):
        messagebox.showinfo("Logs", "No attendance records found.")
        return

    with open(attendance_csv, "r") as f:
        reader = csv.reader(f)
        logs = "\n".join([", ".join(row) for row in reader])
    messagebox.showinfo("Attendance Logs", logs)


def export_logs():
    file_format = selected_format.get()  # Get user choice
    export_path = filedialog.asksaveasfilename(
        defaultextension=f".{file_format}",
        filetypes=[(f"{file_format.upper()} Files", f"*.{file_format}")],
    )

    if export_path:
        if file_format == "xlsx":
            shutil.copy(attendance_xlsx, export_path)
        elif file_format == "csv":
            shutil.copy(attendance_csv, export_path)

        messagebox.showinfo("Export", f"Logs exported to {export_path}")


# GUI setup
root = tk.Tk()
root.title("Face Recognition Attendance")
root.geometry("300x300")

btn_register = tk.Button(root, text="Register Face", command=register_face)
btn_register.pack(pady=10)

btn_recognize = tk.Button(root, text="Check Faces", command=recognize_face)
btn_recognize.pack(pady=10)

btn_logs = tk.Button(root, text="Show Logs", command=show_logs)
btn_logs.pack(pady=10)

btn_export = tk.Button(root, text="Export Logs", command=export_logs)
btn_export.pack(pady=10)

# Variable to store the selected format
selected_format = tk.StringVar()
selected_format.set("xlsx")  # Default format

# Dropdown menu for format selection
format_label = tk.Label(root, text="Select Format:")
format_label.pack(pady=5)

format_menu = tk.OptionMenu(root, selected_format, "csv", "xlsx")
format_menu.pack(pady=5)

load_encodings()
root.mainloop()
