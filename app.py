import os
import pickle
import datetime
import csv
import tkinter as tk
import shutil
from tkinter import (
    messagebox,
    filedialog,
    Toplevel,
    Label,
    Entry,
    Button,
)
from tkinter import ttk

# --- ADDED FOR LOGO ---
from PIL import Image, ImageTk

# --------------------
from openpyxl import Workbook, load_workbook
import cv2
import face_recognition
import numpy as np


# --- SETUP ---
data_dir = "data"
encoding_file = os.path.join(data_dir, "encodings.pkl")
attendance_csv = os.path.join(data_dir, "attendance.csv")
attendance_xlsx = os.path.join(data_dir, "attendance.xlsx")
os.makedirs(data_dir, exist_ok=True)

# Global variables
known_encodings = []
known_person_data = []


# --- CORE FUNCTIONS ---
def save_encodings():
    """Saves the known face encodings and person data to a pickle file."""
    with open(encoding_file, "wb") as f:
        pickle.dump((known_encodings, known_person_data), f)


# def load_encodings():
#     """Loads face encodings and person data from a pickle file."""
#     global known_encodings, known_person_data
#     if os.path.exists(encoding_file):
#         with open(encoding_file, "rb") as f:
#             known_encodings, known_person_data = pickle.load(f)
#     else:
#         known_encodings = []
#         known_person_data = []


# --- CORRECTED GLOBAL FIX ---
def load_encodings():
    """
    Loads face encodings and person data from a pickle file.
    This version also sanitizes the data to remove corrupted/invalid entries,
    fixing the root cause of the "string indices must be integers" error.
    """
    global known_encodings, known_person_data

    # Always start with fresh lists
    clean_encodings = []
    clean_person_data = []

    if os.path.exists(encoding_file):
        try:
            with open(encoding_file, "rb") as f:
                # Load the potentially corrupted data from the file
                loaded_encodings, loaded_person_data = pickle.load(f)

            # Iterate through the loaded data and keep only the valid pairs
            for i, person in enumerate(loaded_person_data):
                # We only keep the item if it's a dictionary
                if isinstance(person, dict):
                    # And if its corresponding encoding exists
                    if i < len(loaded_encodings):
                        clean_person_data.append(person)
                        clean_encodings.append(loaded_encodings[i])

            # If we found and removed bad data, let the user know and save the clean file
            if len(clean_person_data) != len(loaded_person_data):
                print(
                    f"Warning: Corrupted data was found and removed from {encoding_file}."
                )
                # Overwrite the old file with the cleaned data
                with open(encoding_file, "wb") as f:
                    pickle.dump((clean_encodings, clean_person_data), f)

        except Exception as e:
            # This handles cases where the file is completely unreadable
            print(f"Error reading {encoding_file}: {e}. A new file will be used.")
            # The lists will remain empty, which is safe

    # Set the global variables to the cleaned lists
    known_encodings = clean_encodings
    known_person_data = clean_person_data


def mark_attendance(person_data):
    """Marks attendance with NIS, Name, and Class."""
    now = datetime.datetime.now()
    date_str = now.strftime("%Y-%m-%d")
    time_str = now.strftime("%H:%M:%S")
    file_format = selected_format.get()

    row_data = [
        person_data["nis"],
        person_data["name"],
        person_data["class"],
        date_str,
        time_str,
    ]
    header = ["NIS", "Name", "Class", "Date", "Time"]

    if file_format == "xlsx":
        if os.path.exists(attendance_xlsx):
            wb = load_workbook(attendance_xlsx)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(header)

        ws.append(row_data)
        wb.save(attendance_xlsx)

    elif file_format == "csv":
        with open(attendance_csv, "a", newline="") as f:
            writer = csv.writer(f)
            if os.stat(attendance_csv).st_size == 0:
                writer.writerow(header)
            writer.writerow(row_data)


def register_face():
    """Opens a form to get student details before capturing face."""
    form_window = Toplevel(root)
    form_window.title("Register New Student")
    form_window.geometry("300x200")
    form_window.grab_set()

    Label(form_window, text="Name:").pack(pady=(10, 0))
    name_entry = Entry(form_window, width=30)
    name_entry.pack()

    Label(form_window, text="NIS (Student ID):").pack(pady=(10, 0))
    nis_entry = Entry(form_window, width=30)
    nis_entry.pack()

    Label(form_window, text="Class:").pack(pady=(10, 0))
    class_entry = Entry(form_window, width=30)
    class_entry.pack()

    def submit_and_capture():
        name = name_entry.get().strip()
        nis = nis_entry.get().strip()
        class_ = class_entry.get().strip()

        if not name or not nis or not class_:
            messagebox.showerror(
                "Error", "All fields are required.", parent=form_window
            )
            return

        form_window.destroy()
        capture_face_for_registration({"name": name, "nis": nis, "class": class_})

    Button(
        form_window, text="Submit and Capture Face", command=submit_and_capture
    ).pack(pady=20)


def capture_face_for_registration(person_data):
    """Handles the face capture process after getting user details."""
    cap = cv2.VideoCapture(0)
    if not cap.isOpened():
        messagebox.showerror("Error", "Could not open webcam.")
        return

    messagebox.showinfo(
        "Capture",
        f"Capturing for {person_data['name']}.\nLook at the camera and press 's' to capture, or 'q' to cancel.",
    )

    while True:
        ret, frame = cap.read()
        if not ret:
            break
        cv2.imshow("Register Face - Press 's' to save, 'q' to quit", frame)
        key = cv2.waitKey(1) & 0xFF

        if key == ord("s"):
            rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            boxes = face_recognition.face_locations(rgb_frame)
            encodings = face_recognition.face_encodings(rgb_frame, boxes)

            if encodings:
                known_encodings.append(encodings[0])
                known_person_data.append(person_data)
                save_encodings()
                messagebox.showinfo(
                    "Success", f"Face registered for {person_data['name']}!"
                )
            else:
                messagebox.showerror("Error", "No face detected. Please try again.")
            break
        elif key == ord("q"):
            break

    cap.release()
    cv2.destroyAllWindows()


def recognize_face():
    """Recognizes faces using the webcam and marks attendance."""
    load_encodings()
    cap = cv2.VideoCapture(0)
    if not cap.isOpened():
        messagebox.showerror("Error", "Could not open webcam.")
        return

    todays_attendance_nis = set()
    file_format = selected_format.get()
    date_str = datetime.datetime.now().strftime("%Y-%m-%d")

    try:
        if file_format == "xlsx" and os.path.exists(attendance_xlsx):
            wb = load_workbook(attendance_xlsx)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if len(row) > 3 and row[3] == date_str:
                    todays_attendance_nis.add(row[0])
        elif file_format == "csv" and os.path.exists(attendance_csv):
            with open(attendance_csv, "r") as f:
                reader = csv.reader(f)
                next(reader, None)
                for row in reader:
                    if len(row) > 3 and row[3] == date_str:
                        todays_attendance_nis.add(row[0])
    except Exception as e:
        messagebox.showerror("Error", f"Could not read attendance file: {e}")

    while True:
        ret, frame = cap.read()
        if not ret:
            break

        small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
        rgb_small_frame = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB)
        face_locations = face_recognition.face_locations(rgb_small_frame)
        face_encodings = face_recognition.face_encodings(
            rgb_small_frame, face_locations
        )
        tolerance = tolerance_var.get()

        for (top, right, bottom, left), face_encoding in zip(
            face_locations, face_encodings
        ):
            matches = face_recognition.compare_faces(
                known_encodings, face_encoding, tolerance=tolerance
            )
            display_name = "Unknown"
            confidence_str = ""

            if True in matches:
                face_distances = face_recognition.face_distance(
                    known_encodings, face_encoding
                )
                best_match_index = np.argmin(face_distances)
                if matches[best_match_index]:
                    person_data = known_person_data[best_match_index]
                    display_name = person_data["name"]

                    confidence = 1 - face_distances[best_match_index]
                    confidence_str = f"{confidence:.0%}"

                    if person_data["nis"] not in todays_attendance_nis:
                        mark_attendance(person_data)
                        todays_attendance_nis.add(person_data["nis"])
                        print(
                            f"Attendance marked for {display_name} ({person_data['nis']})"
                        )

            top, right, bottom, left = top * 4, right * 4, bottom * 4, left * 4
            cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 0), 2)
            label = f"{display_name} {confidence_str}"
            cv2.putText(
                frame,
                label,
                (left, top - 10),
                cv2.FONT_HERSHEY_SIMPLEX,
                0.7,
                (0, 255, 0),
                2,
            )

        cv2.imshow("Face Recognition - Press 'q' to Exit", frame)
        if cv2.waitKey(1) & 0xFF == ord("q"):
            break

    cap.release()
    cv2.destroyAllWindows()


def show_logs():
    """Shows logs in a new, scrollable window using ttk.Treeview."""
    log_window = Toplevel(root)
    log_window.title("Attendance Logs")
    log_window.geometry("800x500")

    tree_frame = ttk.Frame(log_window, padding="10")
    tree_frame.pack(expand=True, fill="both")

    columns = ("nis", "name", "class", "date", "time")
    tree = ttk.Treeview(tree_frame, columns=columns, show="headings")

    tree.heading("nis", text="NIS")
    tree.heading("name", text="Name")
    tree.heading("class", text="Class")
    tree.heading("date", text="Date")
    tree.heading("time", text="Time")

    tree.column("nis", width=100, anchor=tk.CENTER)
    tree.column("name", width=200)
    tree.column("class", width=100, anchor=tk.CENTER)
    tree.column("date", width=120, anchor=tk.CENTER)
    tree.column("time", width=100, anchor=tk.CENTER)

    scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
    tree.configure(yscrollcommand=scrollbar.set)

    tree.grid(row=0, column=0, sticky="nsew")
    scrollbar.grid(row=0, column=1, sticky="ns")

    tree_frame.grid_rowconfigure(0, weight=1)
    tree_frame.grid_columnconfigure(0, weight=1)

    file_format = selected_format.get()
    log_file_path = attendance_xlsx if file_format == "xlsx" else attendance_csv

    if not os.path.exists(log_file_path):
        tree.insert("", tk.END, values=("No attendance records found.", "", "", "", ""))
        return

    try:
        content = []
        if file_format == "xlsx":
            wb = load_workbook(log_file_path)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                content.append([str(cell) if cell is not None else "" for cell in row])
        else:
            with open(log_file_path, "r", newline="") as f:
                reader = csv.reader(f)
                for row in reader:
                    content.append(row)

        if content:
            for row_data in content[1:]:
                if len(row_data) >= 5:
                    tree.insert("", tk.END, values=row_data[:5])
    except Exception as e:
        messagebox.showerror(
            "Error", f"Failed to read log file: {e}", parent=log_window
        )


# --- CORRECTED FUNCTION ---
def show_registered_students():
    """
    Shows all registered students in a new window using ttk.Treeview.
    This version safely handles corrupted data in the pickle file.
    """
    reg_window = Toplevel(root)
    reg_window.title("Registered Students")
    reg_window.geometry("600x400")
    reg_window.grab_set()

    tree_frame = ttk.Frame(reg_window, padding="10")
    tree_frame.pack(expand=True, fill="both")

    columns = ("nis", "name", "class")
    tree = ttk.Treeview(tree_frame, columns=columns, show="headings")

    tree.heading("nis", text="NIS (Student ID)")
    tree.heading("name", text="Name")
    tree.heading("class", text="Class")

    tree.column("nis", width=120, anchor=tk.CENTER)
    tree.column("name", width=250)
    tree.column("class", width=100, anchor=tk.CENTER)

    scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
    tree.configure(yscrollcommand=scrollbar.set)

    tree.grid(row=0, column=0, sticky="nsew")
    scrollbar.grid(row=0, column=1, sticky="ns")
    tree_frame.grid_rowconfigure(0, weight=1)
    tree_frame.grid_columnconfigure(0, weight=1)

    # --- FIX IS HERE ---
    # Filter the data to only include valid dictionaries, preventing crashes
    valid_data = [item for item in known_person_data if isinstance(item, dict)]

    if not valid_data:
        tree.insert("", tk.END, values=("No students registered yet.", "", ""))
    else:
        # Sort using .get() for safety, in case a key is missing
        sorted_data = sorted(valid_data, key=lambda item: item.get("name", ""))
        for person in sorted_data:
            # Use .get() again for safe access when populating the tree
            nis = person.get("nis", "")
            name = person.get("name", "")
            class_ = person.get("class", "")
            tree.insert("", tk.END, values=(nis, name, class_))


def export_logs():
    """Exports logs to a user-selected location."""
    file_format = selected_format.get()
    source_file = attendance_xlsx if file_format == "xlsx" else attendance_csv

    if not os.path.exists(source_file):
        messagebox.showwarning("Warning", "No attendance file to export.")
        return

    export_path = filedialog.asksaveasfilename(
        defaultextension=f".{file_format}",
        filetypes=[(f"{file_format.upper()} Files", f"*.{file_format}")],
    )

    if export_path:
        try:
            shutil.copy(source_file, export_path)
            messagebox.showinfo(
                "Export", f"Logs successfully exported to {export_path}"
            )
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export file: {e}")


# --- GUI SETUP ---
root = tk.Tk()
root.title("Face Recognition Attendance")
# Adjusted height to make space for the new button
root.geometry("450x680")
root.configure(bg="#f0f0f0")

main_frame = tk.Frame(root, bg="#f0f0f0", padx=20, pady=20)
main_frame.pack(expand=True, fill="both")

### --- ADDED FOR LOGO --- ###
try:
    logo_image = Image.open("logos//logo_right_new.png")
    logo_image = logo_image.resize((160, 154), Image.Resampling.LANCZOS)
    logo_photo = ImageTk.PhotoImage(logo_image)
    logo_label = tk.Label(main_frame, image=logo_photo, bg="#f0f0f0")
    logo_label.image = logo_photo
    logo_label.pack(pady=(0, 10))
except FileNotFoundError:
    print("Warning: logo.png not found. The application will run without a logo.")
except Exception as e:
    print(f"An error occurred while loading the logo: {e}")
### ---------------------- ###


title_label = tk.Label(
    main_frame,
    text="Absensi Sholat Duhur\nSiswa SMAN 1 Paguyangan",
    font=("Helvetica", 16, "bold"),
    bg="#f0f0f0",
    justify=tk.CENTER,
)
title_label.pack(pady=(5, 20))

btn_register = tk.Button(
    main_frame, text="Register New Student", command=register_face, width=30, height=2
)
btn_register.pack(pady=5)

# --- NEW BUTTON ADDED HERE ---
btn_show_registered = tk.Button(
    main_frame,
    text="Show Registered Students",
    command=show_registered_students,
    width=30,
    height=2,
)
btn_show_registered.pack(pady=5)
# -----------------------------

btn_recognize = tk.Button(
    main_frame,
    text="Start Recognition & Attendance",
    command=recognize_face,
    width=30,
    height=2,
)
btn_recognize.pack(pady=5)

btn_logs = tk.Button(
    main_frame, text="Show Attendance Logs", command=show_logs, width=30, height=2
)
btn_logs.pack(pady=5)

btn_export = tk.Button(
    main_frame, text="Export Logs", command=export_logs, width=30, height=2
)
btn_export.pack(pady=5)

format_frame = tk.Frame(main_frame, bg="#f0f0f0")
format_frame.pack(pady=10)
format_label = tk.Label(format_frame, text="Log Format:", bg="#f0f0f0")
format_label.pack(side="left", padx=(0, 10))
selected_format = tk.StringVar(value="xlsx")
format_menu = tk.OptionMenu(format_frame, selected_format, "xlsx", "csv")
format_menu.config(width=8)
format_menu.pack(side="left")

tolerance_frame = tk.Frame(main_frame, bg="#f0f0f0")
tolerance_frame.pack(pady=10, fill="x")
tolerance_label = tk.Label(
    tolerance_frame, text="Recognition Tolerance (Stricter <-> Looser)", bg="#f0f0f0"
)
tolerance_label.pack()
tolerance_var = tk.DoubleVar(value=0.6)
tolerance_slider = tk.Scale(
    tolerance_frame,
    from_=0.4,
    to=0.7,
    resolution=0.05,
    orient="horizontal",
    variable=tolerance_var,
    bg="#f0f0f0",
)
tolerance_slider.pack(fill="x", expand=True)

### --- ADDED FOR DEVELOPER CREDIT --- ###
credit_text = (
    f"Developed by Tim Riset SMAN 1 Paguyangan © {datetime.datetime.now().year}"
)
credit_label = tk.Label(
    main_frame,
    text=credit_text,
    font=("Helvetica", 8),
    fg="gray",
    bg="#f0f0f0",
)
credit_label.pack(side="bottom", pady=5)
### ------------------------------------ ###

# Load existing data when the application starts
load_encodings()
root.mainloop()
